"""每日確認前一天資料是否已更新，結果寄信通知"""
import os, smtplib, glob
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl

CONFIG       = r"C:\temp\parking\config.txt"
GMAIL_CONFIG = r"C:\temp\parking\gmail_config.txt"

with open(CONFIG, encoding="utf-8") as f:
    WORKBOOK = f.read().strip()

MONTHLY_DIR = os.path.join(os.path.dirname(WORKBOOK), "月份資料")
MONTHS_NUM  = ["01","02","03","04","05","06","07","08","09","10","11","12"]

def _monthly_path(date_obj):
    return os.path.join(MONTHLY_DIR, f"{date_obj.year}年{MONTHS_NUM[date_obj.month-1]}月.xlsx")

def check(date_obj):
    """檢查指定日期的資料是否存在，回傳 (ok, total, records)"""
    sheet_name = date_obj.strftime("%Y-%m-%d")
    mpath = _monthly_path(date_obj)

    if not os.path.exists(mpath):
        return False, 0, 0

    try:
        wb = openpyxl.load_workbook(mpath, data_only=True)
    except Exception:
        return False, 0, 0

    if sheet_name not in wb.sheetnames:
        return False, 0, 0

    ws = wb[sheet_name]
    # 找合計列（B欄="合計"）
    total, records = 0, 0
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == "合計":
            v = ws.cell(r, 4).value
            if isinstance(v, (int, float)):
                total = int(v)
            break
        if ws.cell(r, 1).value is not None:
            records += 1

    return True, total, max(records - 1, 0)  # 扣掉表頭

def send_email(subject, body):
    with open(GMAIL_CONFIG, encoding="utf-8-sig") as f:
        lines = f.read().splitlines()
    sender   = lines[0].strip()
    password = lines[1].strip().replace(" ", "")
    receiver = "mikulai1114@gmail.com"

    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = receiver
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(sender, password)
        s.send_message(msg)

def main():
    yesterday = (datetime.now() - timedelta(days=1)).date()
    date_str  = yesterday.strftime("%Y-%m-%d")

    ok, total, records = check(yesterday)

    if ok:
        status  = "[OK] 已更新"
        subject = f"[停車場] {date_str} 資料已更新"
        body    = (
            f"磐鈺雲華商業停車場 每日確認\n\n"
            f"日期：{date_str}\n"
            f"狀態：已更新\n"
            f"筆數：{records} 筆\n"
            f"合計：${total:,}\n"
        )
    else:
        status  = "[缺漏]"
        subject = f"[停車場] {date_str} 資料缺漏，請確認"
        body    = (
            f"磐鈺雲華商業停車場 每日確認\n\n"
            f"日期：{date_str}\n"
            f"狀態：資料未更新\n\n"
            f"請手動執行：\n"
            f"python C:\\temp\\parking\\import.py {date_str}\n"
        )

    print(f"  [每日確認] {date_str} {status}  合計 ${total:,}")

    try:
        send_email(subject, body)
        print(f"  [每日確認] 通知信已寄出")
    except Exception as e:
        print(f"  [每日確認] 寄信失敗：{e}")

if __name__ == "__main__":
    main()
