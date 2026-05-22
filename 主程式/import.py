import sys, os, shutil


from datetime import datetime, timedelta
from playwright.sync_api import sync_playwright
import openpyxl

LOGIN_URL = "http://125.227.64.94:48140/Account/Login"
USERNAME  = "23208598"
PASSWORD  = "23208598"
DL_DIR    = r"C:\temp\parking"
CONFIG    = r"C:\temp\parking\config.txt"
TMP_PATH  = r"C:\temp\parking\wb_save.xlsx"

# 電子發票平台設定
INV_LOGIN_URL = "http://www.tweisc.com/fam/Login.aspx"
INV_IV010_URL = "http://www.tweisc.com/fam/Invoice/IV010.aspx"
INV_FACTORY   = "87164401"
INV_ACCOUNT   = "87164401"
INV_PASSWORD  = "87164401"
INV_POS       = "02"
INV_PERIOD_MAP = {1:'02',2:'02',3:'04',4:'04',5:'06',6:'06',
                  7:'08',8:'08',9:'10',10:'10',11:'12',12:'12'}

MONTHS = ["一月","二月","三月","四月","五月",
          "六月","七月","八月","九月","十月",
          "十一月","十二月"]
MONTHS_NUM = ["01","02","03","04","05","06",
              "07","08","09","10","11","12"]
WEEKDAY_ZH = ["一","二","三","四","五","六","日"]

with open(CONFIG, encoding="utf-8") as _f:
    WORKBOOK = _f.read().strip()

# 月份 Excel 存放資料夾（與主工作簿同層）
MONTHLY_DIR        = os.path.join(os.path.dirname(WORKBOOK), "月份資料")
PNG_DIR            = r"C:\Users\gjsky\OneDrive\桌面\每日圖檔"
VOUCHER_STATS_DIR  = r"C:\Users\gjsky\OneDrive\桌面\磐鈺雲華商業停車場\每日停車券統計"
VOUCHER_STATS_FILE = os.path.join(VOUCHER_STATS_DIR, "停車券統計.xlsx")

def _voucher_monthly_path(date_obj):
    return os.path.join(VOUCHER_STATS_DIR, f"停車券{date_obj.year}年{MONTHS_NUM[date_obj.month-1]}月.xlsx")

DAY_ZH = ["一","二","三","四","五","六","七","八","九","十",
           "十一","十二","十三","十四","十五","十六","十七","十八","十九","二十",
           "二一","二二","二三","二四","二五","二六","二七","二八","二九","三十","三一"]


def _parse_source(excel_path):
    src = openpyxl.load_workbook(excel_path)
    ws  = src.active
    headers = [str(c.value) if c.value else "" for c in ws[1]]

    def ci(names):
        for n in names:
            if n in headers: return headers.index(n)
        return None

    i_plate   = ci(["車牌號碼","車牌"])
    i_entry   = ci(["入場時間","入場"])
    i_exit    = ci(["出場時間","出場"])
    i_amount  = ci(["支付金額","金額"])
    i_method  = ci(["支付方式"])
    i_invoice = ci(["發票號碼"])
    i_carrier = ci(["統編/載具","載具"])

    rows_data, total = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row): continue
        r = list(row)
        plate_val = str(r[i_plate]).strip() if i_plate is not None and r[i_plate] else ""
        if plate_val in ("總計", "合計", "小計", ""):
            continue
        try:
            amt = float(r[i_amount]) if i_amount is not None and r[i_amount] else 0
        except:
            amt = 0
        total += amt
        rows_data.append(r)

    cols = (i_plate, i_entry, i_exit, i_amount, i_method, i_invoice, i_carrier)
    return rows_data, total, cols


def _monthly_excel_path(date_obj):
    """月份 Excel 檔案路徑"""
    fname = f"{date_obj.year}年{MONTHS_NUM[date_obj.month-1]}月.xlsx"
    return os.path.join(MONTHLY_DIR, fname)


def _get_or_create_monthly_wb(monthly_wbs, date_obj):
    """取得或建立月份工作簿（in-memory）"""
    key = (date_obj.year, date_obj.month)
    if key not in monthly_wbs:
        path = _monthly_excel_path(date_obj)
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            # 移除舊版的單一合併工作表（如果存在）
            month_title = MONTHS[date_obj.month - 1]
            if month_title in wb.sheetnames and len(wb.sheetnames) == 1:
                del wb[month_title]
        else:
            wb = openpyxl.Workbook()
            # 移除預設空白工作表
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        monthly_wbs[key] = wb
    return monthly_wbs[key]


def _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols):
    """每天建立獨立分頁，分頁名稱為日期"""
    from openpyxl.styles import Alignment
    wb       = _get_or_create_monthly_wb(monthly_wbs, date_obj)
    sheet_name = date_obj.strftime("%Y-%m-%d")

    # 若已有同名分頁則先刪除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["G"].width = 20  # 載具欄預設寬度

    # 表頭（7欄，不含日期欄）
    center = Alignment(horizontal="center", vertical="center")
    for ci2, h in enumerate(
        ["車牌號碼","進場時間","出場時間","支付金額","支付方式","發票號碼","統編/載具"],
        start=1
    ):
        cell = ws.cell(1, ci2)
        cell.value = h
        cell.alignment = center

    # 寫入資料
    (i_plate, i_entry, i_exit, i_amount, i_method, i_invoice, i_carrier) = cols
    src_cols = [i_plate, i_entry, i_exit, i_amount, i_method, i_invoice, i_carrier]

    # 過濾金額為 0 的列
    def _to_int(v):
        try:
            return int(float(str(v).replace(',', ''))) if v is not None and str(v).strip() else 0
        except (ValueError, TypeError):
            return 0
    rows_data = [rd for rd in rows_data if i_amount is None or _to_int(rd[i_amount]) != 0]

    acct_fmt = '"$"#,##0'
    for ri, rd in enumerate(rows_data, start=2):
        for ci2, sc in enumerate(src_cols, start=1):
            if sc is not None:
                cell = ws.cell(ri, ci2)
                if ci2 == 4:
                    try:
                        cell.value = int(float(str(rd[sc]).replace(',', ''))) if rd[sc] is not None and str(rd[sc]).strip() else 0
                    except (ValueError, TypeError):
                        cell.value = rd[sc]
                    cell.number_format = acct_fmt
                else:
                    cell.value = rd[sc]

    # 合計列在最下方
    from openpyxl.styles import Font
    total_row = len(rows_data) + 2
    total_font = Font(bold=True, color="FF0000")
    for c in [ws.cell(total_row, 2), ws.cell(total_row, 3), ws.cell(total_row, 4)]:
        c.font      = total_font
        c.alignment = center
    ws.cell(total_row, 2).value = "合計"
    ws.cell(total_row, 3).value = date_obj.strftime("%Y/%m/%d")
    amt_cell = ws.cell(total_row, 4)
    amt_cell.value         = int(total)
    amt_cell.number_format = acct_fmt


def _display_width(text):
    """計算顯示寬度：中文字符算 2，英數算 1"""
    w = 0
    for ch in str(text):
        w += 2 if ord(ch) > 127 else 1
    return w


def _auto_col_width(ws):
    """依內容自動調整欄寬（含中文寬度）並加框線（有內容的列整列補齊）"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {}
    max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                w = _display_width(cell.value)
                if w > col_widths.get(cell.column, 0):
                    col_widths[cell.column] = w
                if cell.column > max_col:
                    max_col = cell.column

    # 有內容的列，從第 1 欄到 max_col 整列補齊框線
    if max_col:
        for row in ws.iter_rows(max_col=max_col):
            if any(cell.value is not None for cell in row):
                for cell in row:
                    cell.border = border

    for col, width in col_widths.items():
        col_letter = get_column_letter(col)
        current = ws.column_dimensions[col_letter].width or 0
        ws.column_dimensions[col_letter].width = min(max(width + 2, current), 45)


def _build_stats_sheet(wb, year, month):
    """建立年度收入統計分頁：直接從主工作簿月份分頁讀取每日金額（放置於最前）"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sheet_name = "年度收入統計"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    # 從月份 Excel 每日分頁（YYYY-MM-DD）最下方合計列讀取金額（停車+發票合計）
    monthly_amt = {}
    monthly_sum = {}
    for mi in range(1, 13):
        # 若是當月，直接用傳入的記憶體工作簿（尚未存檔）
        if mi == month:
            m_wb = wb
        else:
            fname = f"{year}年{MONTHS_NUM[mi-1]}月.xlsx"
            fpath = os.path.join(MONTHLY_DIR, fname)
            if not os.path.exists(fpath):
                continue
            try:
                m_wb = openpyxl.load_workbook(fpath, data_only=True)
            except Exception:
                continue
        day_map = {}
        for sname in m_wb.sheetnames:
            # 每日分頁名稱格式：YYYY-MM-DD
            try:
                dt = datetime.strptime(sname, "%Y-%m-%d")
            except ValueError:
                continue
            if dt.year != year or dt.month != mi:
                continue
            dws = m_wb[sname]
            # 從最底列往上找最後一個 B="合計" 的列
            grand = None
            for r in range(dws.max_row, 0, -1):
                if dws.cell(r, 2).value == "合計":
                    v = dws.cell(r, 4).value
                    if isinstance(v, (int, float)):
                        grand = v
                    break
            day_map[dt.day] = grand if grand is not None else ""
        monthly_amt[mi] = day_map
        monthly_sum[mi] = sum(v for v in day_map.values() if isinstance(v, (int, float)))

    # ── 樣式 ──────────────────────────────────────────────
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill  = PatternFill("solid", fgColor="1F4E79")
    title_font  = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subtot_fill = PatternFill("solid", fgColor="BDD7EE")
    subtot_font = Font(bold=True, size=11)
    grand_fill  = PatternFill("solid", fgColor="1F4E79")
    grand_font  = Font(bold=True, color="FFFFFF", size=11)
    weekend_fill = PatternFill("solid", fgColor="D6E4F0")
    center      = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right",  vertical="center")

    num_months  = 12
    last_col    = num_months + 1   # A=日期, B-M=1-12月, N=（備用）

    # ── 第1列：標題 ───────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(num_months + 1)}1")
    c1 = ws["A1"]
    c1.value     = f"磐鈺雲華商業停車場 {year}年度 統計表"
    c1.font      = title_font
    c1.fill      = title_fill
    c1.alignment = center
    ws.row_dimensions[1].height = 30

    # ── 第2列：表頭 ───────────────────────────────────────
    ws.cell(2, 1).value = "日期"
    for mi in range(1, 13):
        ws.cell(2, mi + 1).value = MONTHS[mi - 1]
    for c in range(1, num_months + 2):
        cell = ws.cell(2, c)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 20

    # ── 特殊連假日期集合（(月, 日) tuple）───────────────────
    SPECIAL_HOLIDAYS = {
        2025: {
            (9,27),(9,28),(9,29),
            (10,4),(10,5),(10,6),
            (10,10),(10,11),(10,12),
            (10,24),(10,25),(10,26),
            (12,25),
        },
        2026: {
            (1,1),
            (2,14),(2,15),(2,16),(2,17),(2,18),(2,19),(2,20),(2,21),(2,22),
            (2,27),(2,28),(3,1),
            (4,3),(4,4),(4,5),(4,6),
            (5,1),(5,2),(5,3),
            (6,19),(6,20),(6,21),
            (9,25),(9,26),(9,27),(9,28),
            (10,9),(10,10),(10,11),
            (10,24),(10,25),(10,26),
            (12,25),(12,26),(12,27),
        },
    }
    year_holidays = SPECIAL_HOLIDAYS.get(year, set())

    # ── 第3-33列：每日金額（日1-31）────────────────────────
    for day in range(1, 32):
        row = day + 2
        d_cell = ws.cell(row, 1)
        d_cell.value     = day
        d_cell.alignment = center
        d_cell.border    = border

        for mi in range(1, 13):
            amt = monthly_amt.get(mi, {}).get(day, "")
            cell = ws.cell(row, mi + 1)
            cell.value     = amt
            cell.border    = border
            cell.alignment = right_align if isinstance(amt, (int, float)) else center
            if isinstance(amt, (int, float)):
                cell.number_format = '"$"#,##0'
            white_fill = PatternFill("solid", fgColor="FFFFFF")
            try:
                dt = datetime(year, mi, day)
                if (mi, day) in year_holidays or dt.weekday() >= 5:
                    cell.fill = weekend_fill
                else:
                    cell.fill = white_fill
            except ValueError:
                pass  # 無效日期（如 2/30）

    # ── 小計列 ────────────────────────────────────────────
    subtot_row = 34
    ws.cell(subtot_row, 1).value = "小計"
    for mi in range(1, 13):
        cell = ws.cell(subtot_row, mi + 1)
        cell.value     = monthly_sum.get(mi, "")
        cell.font      = subtot_font
        cell.fill      = subtot_fill
        cell.alignment = right_align
        cell.border    = border
        if isinstance(monthly_sum.get(mi, ""), (int, float)):
            cell.number_format = '"$"#,##0'
    ws.cell(subtot_row, 1).font      = subtot_font
    ws.cell(subtot_row, 1).fill      = subtot_fill
    ws.cell(subtot_row, 1).alignment = center
    ws.cell(subtot_row, 1).border    = border

    # ── 總計列 ────────────────────────────────────────────
    grand_row   = 35
    grand_total = sum(v for v in monthly_sum.values() if isinstance(v, (int, float)))
    ws.merge_cells(f"A{grand_row}:{get_column_letter(num_months + 1)}{grand_row}")
    g_cell = ws.cell(grand_row, 1)
    g_cell.value     = f"年度總計：{grand_total:,} 元"
    g_cell.font      = grand_font
    g_cell.fill      = grand_fill
    g_cell.alignment = center
    g_cell.border    = border

    ws.row_dimensions[grand_row].height = 22

    # ── 欄寬 ──────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7
    for c in range(2, num_months + 2):
        ws.column_dimensions[get_column_letter(c)].width = 12


def _build_nav_sheet(wb):
    """建立或更新「選擇日期」導覽分頁（下拉選單 + 超連結）"""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    nav_name = "選擇日期"
    # 取得所有日期分頁（排除導覽分頁本身）
    date_sheets = [s for s in wb.sheetnames if s != nav_name]

    # 刪除舊導覽分頁，重建置於最前
    if nav_name in wb.sheetnames:
        del wb[nav_name]
    ws = wb.create_sheet(nav_name, 0)

    # 樣式
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    label_font  = Font(bold=True, size=11)
    link_font   = Font(color="0000FF", underline="single", size=11)

    # 標題列
    ws.merge_cells("A1:C1")
    ws["A1"].value     = "請選擇日期後點擊「前往」"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 標籤
    ws["A2"].value = "選擇日期："
    ws["A2"].font  = label_font
    ws["A2"].border = border

    # 下拉選單儲存格
    ws["B2"].value     = date_sheets[0] if date_sheets else ""
    ws["B2"].font      = Font(size=11)
    ws["B2"].border    = border
    ws["B2"].alignment = Alignment(horizontal="center")

    # 資料驗證：下拉清單
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(date_sheets) + '"',
        allow_blank=False,
        showDropDown=False
    )
    ws.add_data_validation(dv)
    dv.add(ws["B2"])

    # 超連結「前往」按鈕（HYPERLINK 公式）
    ws["C2"].value     = '=HYPERLINK("#\'"&B2&"\'!A1","▶ 前往")'
    ws["C2"].font      = link_font
    ws["C2"].border    = border
    ws["C2"].alignment = Alignment(horizontal="center")

    # 下方列出所有日期的快速連結
    ws["A4"].value = "所有日期快速連結："
    ws["A4"].font  = Font(bold=True, size=10)

    for i, sheet in enumerate(date_sheets):
        row = i + 5
        ws.cell(row, 1).value     = f'=HYPERLINK("#\'{sheet}\'!A1","{sheet}")'
        ws.cell(row, 1).font      = link_font
        ws.cell(row, 1).border    = border
        ws.cell(row, 1).alignment = Alignment(horizontal="center")

    # 欄寬
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12


def _save_monthly_wbs(monthly_wbs):
    os.makedirs(MONTHLY_DIR, exist_ok=True)
    tmp = r"C:\temp\parking\month_save.xlsx"
    for (year, month), wb in monthly_wbs.items():
        for ws in wb.worksheets:
            _auto_col_width(ws)
        _build_nav_sheet(wb)          # 選擇日期（第2分頁）
        _build_stats_sheet(wb, year, month)  # 月份收入統計（第1分頁）
        key_obj = datetime(year, month, 1)
        path = _monthly_excel_path(key_obj)
        wb.save(tmp)
        shutil.copy2(tmp, path)


def _write_monthly(wb, month_name, day, total):
    wb[month_name].cell(day + 2, 4).value = int(total)


def _write_daily_summary(wb, date_obj, rows_data, total, i_amount):
    date_str    = date_obj.strftime("%Y/%m/%d")
    weekday_zh  = WEEKDAY_ZH[date_obj.weekday()]
    type_str    = "假日" if date_obj.weekday() >= 5 else "平日"
    total_trips = len(rows_data)
    paid_trips  = sum(1 for rd in rows_data
                      if (float(rd[i_amount]) if i_amount is not None and rd[i_amount] else 0) > 0)
    free_trips  = total_trips - paid_trips
    cash_income = int(total)

    ws_d = wb["每日彙整"]
    target_row = None
    for row_idx in range(2, ws_d.max_row + 1):
        val = ws_d.cell(row_idx, 1).value
        if val is not None and str(val).strip() == date_str:
            target_row = row_idx
            break
    if target_row is None:
        ws_d.insert_rows(2)
        target_row = 2

    ws_d.cell(target_row, 1).value = date_str
    ws_d.cell(target_row, 2).value = weekday_zh
    ws_d.cell(target_row, 3).value = type_str
    ws_d.cell(target_row, 4).value = total_trips
    ws_d.cell(target_row, 5).value = paid_trips
    ws_d.cell(target_row, 6).value = free_trips
    ws_d.cell(target_row, 7).value = cash_income
    ws_d.cell(target_row, 8).value = 0
    ws_d.cell(target_row, 9).value = cash_income
    return total_trips, cash_income


def _save_wb(wb):
    import time
    wb.save(TMP_PATH)
    for attempt in range(6):
        try:
            shutil.copy2(TMP_PATH, WORKBOOK)
            return
        except PermissionError:
            if attempt == 0:
                print("\n  ⚠ 主檔被 Excel 開啟，等待 10 秒後重試...")
            time.sleep(10)
    raise PermissionError("無法寫入主檔，請關閉 Excel 後重試")


def _set_filter(frame, page, start_str, end_str):
    frame.evaluate(f"""
        (function(){{
            var inputs = document.querySelectorAll('input.layui-input, input[type=text]');
            if(inputs[0]){{ inputs[0].value='{start_str}'; inputs[0].dispatchEvent(new Event('change',{{bubbles:true}})); }}
            if(inputs[1]){{ inputs[1].value='{end_str}';   inputs[1].dispatchEvent(new Event('change',{{bubbles:true}})); }}
            var selects = document.querySelectorAll('select');
            for(var sel of selects){{
                for(var opt of sel.options){{
                    if(opt.text && opt.text.includes('現金支付')){{
                        sel.value=opt.value; sel.dispatchEvent(new Event('change',{{bubbles:true}})); break;
                    }}
                }}
            }}
        }})();
    """)
    page.wait_for_timeout(800)


def _search(frame, page):
    frame.evaluate("""
        (function(){
            var btns=document.querySelectorAll('button,.layui-btn');
            for(var btn of btns){
                if((btn.innerText||'').replace(/\\s+/g,'').includes('搜索')){btn.click();break;}
            }
        })();
    """)
    try:
        frame.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    page.wait_for_timeout(2000)


def _export(frame, page, target):
    save_path = os.path.join(DL_DIR, f"parking_{target}.xlsx")
    with page.expect_download(timeout=20000) as dl_info:
        frame.evaluate("""
            (function(){
                var btns=document.querySelectorAll('button,.layui-btn,a');
                for(var btn of btns){
                    if((btn.innerText||'').trim().includes('Excel')){btn.click();break;}
                }
            })();
        """)
    dl_info.value.save_as(save_path)
    return save_path


def _login_and_open(page):
    page.goto(LOGIN_URL, timeout=30000)
    page.fill("input[name=username]", USERNAME)
    page.fill("input[name=password]", PASSWORD)
    page.keyboard.press("Enter")
    page.wait_for_load_state("networkidle", timeout=20000)

    nav = page.locator(".layui-nav-item")
    nav.nth(2).click()
    page.wait_for_timeout(1000)
    nav.nth(2).locator("dd a, .layui-nav-child a").first.click()

    frame = None
    for _ in range(10):
        page.wait_for_timeout(1000)
        for f in page.frames:
            if "ChargeRecord" in f.url:
                frame = f
                break
        if frame:
            break
    if not frame:
        raise Exception("找不到支付流水頁面")
    return frame


# ── 單日匯入 ──────────────────────────────────────────────
def run(target=None):
    if target is None:
        target = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    date_obj   = datetime.strptime(target, "%Y-%m-%d")
    month_name = MONTHS[date_obj.month - 1]
    day        = date_obj.day
    os.makedirs(DL_DIR, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        print(f"[1] 登入 ({target})...")
        frame = _login_and_open(page)
        print(f"    OK")

        print(f"[2] 設定日期 + 現金支付...")
        _set_filter(frame, page, f"{target} 00:00:00", f"{target} 23:59:59")

        print("[3] 搜索...")
        _search(frame, page)

        print("[4] 匯出 Excel...")
        save_path = _export(frame, page, target)

        print(f"[4b] 下載停車券資料（眾陽時數券）...")
        voucher_path = os.path.join(DL_DIR, f"voucher_{target}.xlsx")
        v_count, v_rows = 0, []
        try:
            _set_filter_voucher(frame, page, f"{target} 00:00:00", f"{target} 23:59:59")
            _search(frame, page)
            row_count = frame.evaluate("""
                (function(){
                    var rows = document.querySelectorAll('table tbody tr');
                    if(rows.length === 0) return 0;
                    var txt = (rows[0].innerText || '').trim();
                    if(txt.indexOf('無數據') >= 0 ||
                       txt.indexOf('暫無數據') >= 0 ||
                       txt.indexOf('No data') >= 0) return 0;
                    return rows.length;
                })();
            """)
            if row_count == 0:
                print(f"  停車券：{target} 無資料")
            else:
                with page.expect_download(timeout=20000) as dl_info:
                    frame.evaluate("""
                        (function(){
                            var btns=document.querySelectorAll('button,.layui-btn,a');
                            for(var btn of btns){
                                if((btn.innerText||'').trim().includes('Excel')){btn.click();break;}
                            }
                        })();
                    """)
                dl_info.value.save_as(voucher_path)
                v_count, v_rows = _parse_voucher_excel(voucher_path)
                print(f"  停車券：{v_count}張")
        except Exception as e:
            v_rows = []
            print(f"  停車券下載失敗：{e}")
        browser.close()

    rows_data, total, cols = _parse_source(save_path)
    print(f"  筆數:{len(rows_data)}  合計:${int(total):,}")

    # 主工作簿
    wb = openpyxl.load_workbook(WORKBOOK)
    _write_monthly(wb, month_name, day, total)
    _write_daily_summary(wb, date_obj, rows_data, total, cols[3])
    _save_wb(wb)

    # 月份 Excel
    monthly_wbs = {}
    _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols)
    _save_monthly_wbs(monthly_wbs)

    print(f"  已寫入：{month_name} 第{day}天 = ${int(total):,}")

    # 電子發票
    import time as _time
    print(f"[5] 查詢電子發票 ({target})...")
    with sync_playwright() as p2:
        browser2 = p2.chromium.launch(headless=True)
        ctx2     = browser2.new_context(accept_downloads=True)
        page2    = ctx2.new_page()
        if _inv_login(page2):
            page2.on("dialog", lambda d: d.accept())
            inv_path = _inv_fetch_excel(page2, date_obj)
            if inv_path:
                inv_rows = _inv_parse_excel(inv_path)
                if inv_rows:
                    mpath      = _monthly_excel_path(date_obj)
                    sheet_name = date_obj.strftime("%Y-%m-%d")
                    wb_m       = openpyxl.load_workbook(mpath)
                    if sheet_name in wb_m.sheetnames:
                        ws_m = wb_m[sheet_name]
                        body_vals = []
                        for r in range(max(1, ws_m.max_row - 5), ws_m.max_row + 1):
                            for c in (1, 2):
                                v = ws_m.cell(r, c).value
                                if v:
                                    body_vals.append(v)
                        if "電子發票明細" not in body_vals:
                            _append_invoice_section(ws_m, inv_rows, date_obj)
                            _auto_col_width(ws_m)
                            _build_stats_sheet(wb_m, date_obj.year, date_obj.month)
                            tmp_m = r"C:\temp\parking\month_save.xlsx"
                            wb_m.save(tmp_m)
                            for _ in range(6):
                                try:
                                    shutil.copy2(tmp_m, mpath)
                                    break
                                except PermissionError:
                                    _time.sleep(10)
                            print(f"  發票：{len(inv_rows)} 張已寫入，年度統計已更新")
                        else:
                            print(f"  發票：已存在，略過")
                else:
                    print(f"  發票：查無資料")
            else:
                print(f"  發票：查無資料")
        else:
            print(f"  發票：登入失敗")
        browser2.close()

    _rebuild_invoice_record()
    _update_voucher_stats(date_obj, v_count, v_rows)

    # 截圖（每日固定儲存）
    print("[6] 截圖存檔...")
    mpath = _monthly_excel_path(date_obj)
    sheet_name = date_obj.strftime("%Y-%m-%d")
    os.makedirs(PNG_DIR, exist_ok=True)
    png1 = os.path.join(PNG_DIR, f"daily_{date_obj.strftime('%Y-%m-%d')}.png")
    png2 = os.path.join(PNG_DIR, f"stats_{date_obj.strftime('%Y-%m-%d')}.png")
    png3 = os.path.join(PNG_DIR, f"voucher_{date_obj.strftime('%Y-%m-%d')}.png")
    png4 = os.path.join(PNG_DIR, f"voucher_daily_{date_obj.strftime('%Y-%m-%d')}.png")
    png_ok = False
    try:
        _sheet_to_png(mpath, sheet_name, png1)
        _sheet_to_png(mpath, "年度收入統計", png2)
        print(f"  已儲存：{os.path.basename(png1)}")
        print(f"  已儲存：{os.path.basename(png2)}")
        png_ok = True
    except Exception as e:
        print(f"  [截圖] 失敗：{e}")

    try:
        _sheet_to_png(VOUCHER_STATS_FILE, f"{date_obj.year}年停車券統計", png3)
        print(f"  已儲存：{os.path.basename(png3)}")
    except Exception as e:
        print(f"  [截圖] 停車券統計失敗：{e}")
        png3 = None

    try:
        vpath = _voucher_monthly_path(date_obj)
        _sheet_to_png(vpath, date_obj.strftime("%Y-%m-%d"), png4)
        print(f"  已儲存：{os.path.basename(png4)}")
    except Exception as e:
        print(f"  [截圖] 停車券每日分頁失敗：{e}")
        png4 = None

    # 寄信（需 gmail_config.txt 設定）
    if png_ok:
        try:
            send_list = [png1, png2]
            if png4 and os.path.exists(png4):
                send_list.append(png4)
            if png3 and os.path.exists(png3):
                send_list.append(png3)
            _send_report_email(date_obj, send_list)
        except Exception as e:
            print(f"  [郵件] 寄送失敗：{e}")

    print("完成！")


# ── 批次匯入 ──────────────────────────────────────────────
def run_batch(start_date, end_date):
    d   = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date,   "%Y-%m-%d")
    dates = []
    while d <= end:
        dates.append(d)
        d += timedelta(days=1)

    total_days = len(dates)
    print(f"批次匯入：{start_date} ~ {end_date}，共 {total_days} 天\n")
    os.makedirs(DL_DIR, exist_ok=True)
    os.makedirs(MONTHLY_DIR, exist_ok=True)

    wb          = openpyxl.load_workbook(WORKBOOK)
    monthly_wbs = {}   # (year, month) → workbook
    errors      = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        print("登入中...")
        frame = _login_and_open(page)
        print("登入成功，開始逐日匯入\n")

        for i, date_obj in enumerate(dates):
            target     = date_obj.strftime("%Y-%m-%d")
            month_name = MONTHS[date_obj.month - 1]
            day        = date_obj.day

            print(f"[{i+1:3}/{total_days}] {target}", end="  ", flush=True)

            try:
                _set_filter(frame, page,
                            f"{target} 00:00:00",
                            f"{target} 23:59:59")
                _search(frame, page)
                save_path = _export(frame, page, target)
                rows_data, total, cols = _parse_source(save_path)

                _write_monthly(wb, month_name, day, total)
                _write_daily_summary(wb, date_obj, rows_data, total, cols[3])
                _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols)

                print(f"→ {len(rows_data):3}筆  ${int(total):>7,}")
            except Exception as e:
                msg = str(e)[:60]
                print(f"→ 略過（{msg}）")
                errors.append((target, msg))

            # 每 10 天存一次
            if (i + 1) % 10 == 0:
                _save_wb(wb)
                _save_monthly_wbs(monthly_wbs)
                print(f"         ↳ 已存檔 ({i+1}/{total_days})")

        browser.close()

    _save_wb(wb)
    _save_monthly_wbs(monthly_wbs)
    print(f"\n批次完成！共 {total_days} 天，失敗 {len(errors)} 筆")
    if errors:
        print("失敗清單：")
        for d, e in errors:
            print(f"  {d}: {e}")


def rebuild_monthly():
    """從本機已下載的 Excel 重建月份工作簿（每天一個分頁）"""
    import glob, time
    files = sorted(glob.glob(os.path.join(DL_DIR, "parking_*.xlsx")))
    print(f"找到 {len(files)} 個每日 Excel，開始重建月份工作簿...\n")

    monthly_wbs = {}
    for path in files:
        fname = os.path.basename(path)          # parking_2026-01-01.xlsx
        date_str = fname[8:18]                  # 2026-01-01
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            continue

        rows_data, total, cols = _parse_source(path)
        _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols)

        # 若本機有對應發票檔，一併寫入
        inv_path = os.path.join(DL_DIR, f"invoice_{date_str}.xlsx")
        if os.path.exists(inv_path):
            inv_rows = _inv_parse_excel(inv_path)
            if inv_rows:
                key = (date_obj.year, date_obj.month)
                wb  = monthly_wbs[key]
                ws  = wb[date_str]
                _append_invoice_section(ws, inv_rows, date_obj)
                print(f"  {date_str}  {len(rows_data):3}筆  ${int(total):>7,}  [發票 {len(inv_rows)} 張]")
                continue
        print(f"  {date_str}  {len(rows_data):3}筆  ${int(total):>7,}")

    # 存檔
    os.makedirs(MONTHLY_DIR, exist_ok=True)
    tmp = r"C:\temp\parking\month_save.xlsx"
    for (year, month), wb in monthly_wbs.items():
        for ws in wb.worksheets:
            _auto_col_width(ws)
        _build_nav_sheet(wb)
        _build_stats_sheet(wb, year, month)
        key_obj = datetime(year, month, 1)
        dest = _monthly_excel_path(key_obj)
        wb.save(tmp)
        for attempt in range(6):
            try:
                shutil.copy2(tmp, dest)
                break
            except PermissionError:
                print(f"  [!] {os.path.basename(dest)} 被開啟中，等待 10 秒...")
                time.sleep(10)
        print(f"\n已儲存：{os.path.basename(dest)}（{len(wb.sheetnames)} 個分頁）")

    _rebuild_invoice_record()
    print("\n重建完成！")


# ── 電子發票平台函數 ─────────────────────────────────────────

def _inv_login(page):
    """登入電子發票平台，成功回傳 True"""
    import io, time
    import ddddocr
    from PIL import Image, ImageEnhance
    ocr = ddddocr.DdddOcr(show_ad=False)
    for _ in range(10):
        page.goto(INV_LOGIN_URL, timeout=30000)
        page.wait_for_load_state("networkidle")
        time.sleep(1)
        raw = page.query_selector("#ImageCode_Image1").screenshot()
        img = Image.open(io.BytesIO(raw)).convert("RGB")
        img = img.resize((img.width * 3, img.height * 3), Image.LANCZOS)
        img = ImageEnhance.Contrast(img).enhance(2.0)
        buf = io.BytesIO(); img.save(buf, format="PNG")
        code = ocr.classification(buf.getvalue())
        page.fill("#tbx_FactoryID", INV_FACTORY)
        page.fill("#tbx_No",        INV_ACCOUNT)
        page.fill("#tbx_Pwd",       INV_PASSWORD)
        page.fill("#ImageCode_tbx_Code", code)
        try:
            with page.expect_navigation(timeout=15000, wait_until="networkidle"):
                page.click("#btn_OK")
        except Exception:
            pass
        if "Login" not in page.url:
            return True
        time.sleep(0.5)
    return False


def _inv_fetch_excel(page, date_obj):
    """查詢 IV010，下載當日發票 Excel，回傳本機路徑或 None"""
    import time
    date_str = date_obj.strftime("%Y/%m/%d")
    year_str = str(date_obj.year)
    period   = INV_PERIOD_MAP[date_obj.month]
    save_path = os.path.join(DL_DIR, f"invoice_{date_obj.strftime('%Y-%m-%d')}.xlsx")

    page.goto(INV_IV010_URL, timeout=30000)
    page.wait_for_load_state("networkidle")
    time.sleep(1)

    page.select_option("#ddl_src_YYYY", year_str)
    page.select_option("#ddl_src_MM",   period)
    page.fill("#tbx_src_InvoiceDate",   date_str)
    page.select_option("#ddl_src_Machine", INV_POS)
    page.select_option("#ddl_src_State",   "1")

    page.click("#btn_OK")
    page.wait_for_load_state("networkidle", timeout=30000)
    time.sleep(2)

    if "查無資料" in page.inner_text("body"):
        return None

    try:
        with page.expect_download(timeout=30000) as dl:
            page.click("#btn_SrcExcel")
        dl.value.save_as(save_path)
        return save_path
    except Exception:
        return None


def _inv_parse_excel(path):
    """解析發票 Excel，回傳 [(發票日期, 發票號碼, 買方統編, 發票金額), ...]"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "序號":
                header_found = True
            continue
        # 跳過小計列
        if row[0] is None or str(row[0]).strip() in ("", ) or "總計" in str(row[0]):
            continue
        inv_date  = str(row[3]).strip() if row[3] else ""
        inv_no    = str(row[4]).strip() if row[4] else ""
        buyer_id  = str(row[5]).strip() if row[5] else ""
        inv_amt   = row[10] if isinstance(row[10], (int, float)) else 0
        rows.append((inv_date, inv_no, buyer_id, inv_amt))
    return rows


def _append_invoice_section(ws, invoice_rows, date_obj=None):
    """在每日分頁合計列下方加入電子發票明細區塊（欄位對齊上方停車表格）"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    no_fill    = PatternFill("solid", fgColor="FFFFFF")
    center     = Alignment(horizontal="center", vertical="center")
    acct_fmt   = '"$"#,##0'

    COL_DATE   = 2
    COL_TYPE   = 3
    COL_AMT    = 4
    COL_INV_NO = 6
    COL_BUYER  = 7

    # 資料列（直接從合計列下一列開始，不加表頭）
    inv_top    = ws.max_row + 1  # 發票區塊起始列（供下方讀取停車合計用）
    for ri, (inv_date, inv_no, buyer_id, inv_amt) in enumerate(invoice_rows, start=inv_top):
        ws.cell(ri, COL_DATE).value   = inv_date
        ws.cell(ri, COL_TYPE).value   = "停車券"
        ws.cell(ri, COL_AMT).value    = inv_amt
        ws.cell(ri, COL_AMT).number_format = acct_fmt
        ws.cell(ri, COL_INV_NO).value = inv_no
        ws.cell(ri, COL_BUYER).value  = buyer_id if buyer_id else ""
        for col in (COL_DATE, COL_TYPE, COL_AMT, COL_INV_NO, COL_BUYER):
            ws.cell(ri, col).border    = border
            ws.cell(ri, col).alignment = center
            ws.cell(ri, col).fill      = no_fill

    # 合計列（上方停車合計 + 發票金額合計）
    total_font = Font(bold=True, color="FF0000")
    total_row  = inv_top + len(invoice_rows)
    inv_total  = sum(amt for (_, _, _, amt) in invoice_rows if isinstance(amt, (int, float)))

    # 讀取上方停車表格合計金額（B欄="合計" 的 D欄值）
    parking_total = 0
    for r in range(1, inv_top):
        if ws.cell(r, 2).value == "合計":
            v = ws.cell(r, 4).value
            if isinstance(v, (int, float)):
                parking_total = v
            break

    grand_total = parking_total + inv_total

    for col in (COL_DATE, COL_TYPE, COL_AMT, COL_INV_NO, COL_BUYER):
        c = ws.cell(total_row, col)
        c.border    = border
        c.alignment = center
        c.font      = total_font
        c.fill      = no_fill
    ws.cell(total_row, COL_DATE).value         = "合計"
    if date_obj:
        ws.cell(total_row, COL_TYPE).value     = date_obj.strftime("%Y/%m/%d")
    ws.cell(total_row, COL_AMT).value          = grand_total
    ws.cell(total_row, COL_AMT).number_format  = acct_fmt

    # 分頁標籤標示為黃色（有發票資料）
    ws.sheet_properties.tabColor = "FFFF00"


INVOICE_RECORD = r"C:\temp\parking\電子發票紀錄.xlsx"

def _rebuild_invoice_record():
    """從月份 Excel 重建電子發票紀錄.xlsx"""
    import glob, re
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # 掃描所有月份 Excel，找出有黃色分頁（有發票）的每日分頁
    records = []  # [(date_str, inv_count)]
    xlsx_files = [f for f in sorted(glob.glob(os.path.join(MONTHLY_DIR, "*.xlsx")))
                  if not os.path.basename(f).startswith("~$")]
    for fpath in xlsx_files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception:
            continue
        for sname in wb.sheetnames:
            try:
                datetime.strptime(sname, "%Y-%m-%d")
            except ValueError:
                continue
            ws = wb[sname]
            # 找停車合計列（第一個 B="合計" 的列）
            parking_total_row = None
            for r in range(1, ws.max_row + 1):
                if ws.cell(r, 2).value == "合計":
                    parking_total_row = r
                    break
            if parking_total_row is None:
                continue
            # 停車合計列之後若還有 D 欄數值列即為發票資料
            inv_count = 0
            for r in range(parking_total_row + 1, ws.max_row + 1):
                b = ws.cell(r, 2).value
                d = ws.cell(r, 4).value
                if b != "合計" and isinstance(d, (int, float)):
                    inv_count += 1
            if inv_count > 0:
                records.append((sname, inv_count))

    if not records:
        print("  發票紀錄：無資料")
        return

    # 建立 Excel
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    hdr_fill   = PatternFill("solid", fgColor="2E75B6")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    total_font = Font(bold=True, color="FF0000")
    month_colors = ["DEEAF1","E2EFDA","FFF2CC","FCE4D6","EAD1DC",
                    "D9E1F2","FCE4D6","E2EFDA","D6E4F0","FFF2CC","DEEAF1","EAD1DC"]

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "電子發票紀錄"

    ws2.merge_cells("A1:C1")
    ws2["A1"].value     = f"磐鈺雲華商業停車場 電子發票紀錄"
    ws2["A1"].font      = title_font
    ws2["A1"].fill      = title_fill
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 30

    for col, h in [(1,"月份"), (2,"日期"), (3,"張數")]:
        c = ws2.cell(2, col)
        c.value = h; c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = border

    MONTHS_ZH = ["一","二","三","四","五","六","七","八","九","十","十一","十二"]
    row = 3
    for date_str, cnt in records:
        mo = int(date_str[5:7])
        fill = PatternFill("solid", fgColor=month_colors[mo - 1])
        for col, val in [(1, MONTHS_ZH[mo-1]+"月"), (2, date_str), (3, cnt)]:
            c = ws2.cell(row, col)
            c.value = val; c.fill = fill
            c.alignment = center; c.border = border
            c.font = Font(size=11)
        row += 1

    total = sum(c for _, c in records)
    ws2.merge_cells(f"A{row}:B{row}")
    for col in (1, 2, 3):
        ws2.cell(row, col).border = border
        ws2.cell(row, col).alignment = center
    ws2.cell(row, 1).value = "合計"
    ws2.cell(row, 1).font  = total_font
    ws2.cell(row, 3).value = total
    ws2.cell(row, 3).font  = total_font

    col_widths = {1: 0, 2: 0, 3: 0}
    for row in ws2.iter_rows():
        for cell in row:
            if cell.column in col_widths and cell.value:
                # 中文字每個算2個寬度單位
                val_str = str(cell.value)
                w = sum(2 if ord(c) > 127 else 1 for c in val_str)
                col_widths[cell.column] = max(col_widths[cell.column], w)
    from openpyxl.utils import get_column_letter
    for col, w in col_widths.items():
        ws2.column_dimensions[get_column_letter(col)].width = w + 4

    # 確保三欄合計寬度足以容納標題文字
    title_text = ws2["A1"].value or ""
    title_w = sum(2 if ord(c) > 127 else 1 for c in title_text) + 4
    total_w  = sum(ws2.column_dimensions[get_column_letter(c)].width for c in range(1, 4))
    if total_w < title_w:
        extra = (title_w - total_w) / 3
        for col in range(1, 4):
            ws2.column_dimensions[get_column_letter(col)].width += extra

    wb2.save(INVOICE_RECORD)
    print(f"  發票紀錄已更新：{len(records)} 天，合計 {total} 張")


GMAIL_CONFIG = r"C:\temp\parking\gmail_config.txt"


def _set_filter_voucher(frame, page, start_str, end_str):
    """設定日期篩選（重置支付方式為全部，Python 端再篩出眾陽時數券）"""
    frame.evaluate(f"""
        (function(){{
            var inputs = document.querySelectorAll('input.layui-input, input[type=text]');
            if(inputs[0]){{ inputs[0].value='{start_str}'; inputs[0].dispatchEvent(new Event('change',{{bubbles:true}})); }}
            if(inputs[1]){{ inputs[1].value='{end_str}';   inputs[1].dispatchEvent(new Event('change',{{bubbles:true}})); }}
            // 重置支付方式為全部（不篩選）
            var paySel = document.querySelector('select[name="PayType"]');
            if(paySel){{ paySel.value='0'; paySel.dispatchEvent(new Event('change',{{bubbles:true}})); }}
        }})();
    """)
    page.wait_for_timeout(800)


def _parse_voucher_excel(path):
    """解析停車券 Excel，回傳 (張數, 資料列清單)"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    def ci(names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h: return i
        return None

    i_plate    = ci(["車牌"])
    i_entry    = ci(["入場", "進場"])
    i_exit     = ci(["出場"])
    i_amount   = ci(["支付金額", "金額"])
    i_discount = ci(["電子優惠", "子優惠", "優惠"])
    i_method   = ci(["支付方式"])

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue
        plate = str(row[i_plate]).strip() if i_plate is not None and row[i_plate] else ""
        if not plate or plate in ("總計", "合計", "小計"):
            continue
        # 只保留眾陽時數券（支付方式含「眾陽」或「時數」，或電子優惠 > 0）
        method = str(row[i_method]).strip() if i_method is not None and row[i_method] else ""
        disc   = row[i_discount] if i_discount is not None else 0
        try:
            disc_val = float(disc) if disc else 0
        except (ValueError, TypeError):
            disc_val = 0
        if "眾陽" not in method and "時數" not in method and disc_val <= 0:
            continue
        rows.append((
            plate,
            row[i_entry]    if i_entry    is not None else "",
            row[i_exit]     if i_exit     is not None else "",
            row[i_amount]   if i_amount   is not None else 0,
            row[i_discount] if i_discount is not None else 0,
            row[i_method]   if i_method   is not None else "",
        ))
    return len(rows), rows


def _create_voucher_stats_wb(year):
    """建立新的折扣券年度統計工作簿"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    title_fill = PatternFill("solid", fgColor="FFFFFF")
    title_font = Font(bold=True, color="000000", size=14)
    hdr_fill   = PatternFill("solid", fgColor="D9D9D9")
    hdr_font   = Font(bold=True, color="000000", size=11)
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    total_fill = PatternFill("solid", fgColor="D9D9D9")
    total_font = Font(bold=True, color="000000", size=11)
    grand_fill = PatternFill("solid", fgColor="D9D9D9")
    grand_font = Font(bold=True, color="000000", size=11)

    # 特殊連假日期集合（同現金支付年度統計）
    SPECIAL_HOLIDAYS = {
        2025: {
            (9,27),(9,28),(9,29),
            (10,4),(10,5),(10,6),
            (10,10),(10,11),(10,12),
            (10,24),(10,25),(10,26),
            (12,25),
        },
        2026: {
            (1,1),
            (2,14),(2,15),(2,16),(2,17),(2,18),(2,19),(2,20),(2,21),(2,22),
            (2,27),(2,28),(3,1),
            (4,3),(4,4),(4,5),(4,6),
            (5,1),(5,2),(5,3),
            (6,19),(6,20),(6,21),
            (9,25),(9,26),(9,27),(9,28),
            (10,9),(10,10),(10,11),
            (10,24),(10,25),(10,26),
            (12,25),(12,26),(12,27),
        },
    }
    year_holidays = SPECIAL_HOLIDAYS.get(year, set())

    def _day_fill(month, day):
        """假日/特殊連假→綠色，其他→白色"""
        from datetime import date as _date
        import calendar
        try:
            dt = _date(year, month, day)
            if dt.weekday() >= 5 or (month, day) in year_holidays:
                return green_fill
        except ValueError:
            pass
        return white_fill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年停車券統計"

    # 標題列（先設框線再合併，確保整列上下框線可見）
    for c in range(1, 14):
        cell = ws.cell(1, c)
        cell.fill = title_fill; cell.font = title_font
        cell.alignment = center; cell.border = border
    ws.merge_cells("A1:M1")
    ws["A1"].value = f"磐鈺雲華商業停車場停車券(已使用張數)統計表"
    ws.row_dimensions[1].height = 30

    # 表頭
    ws.cell(2, 1).value = "月/日"
    for m in range(1, 13):
        ws.cell(2, m + 1).value = f"{m}月"
    for c in range(1, 14):
        cell = ws.cell(2, c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[2].height = 20

    # 日期列（31天）：日期欄白色，資料欄依假日判斷
    for day in range(1, 32):
        row = day + 2
        ws.cell(row, 1).value     = DAY_ZH[day - 1]
        ws.cell(row, 1).font      = Font(bold=True, size=11)
        ws.cell(row, 1).fill      = white_fill
        ws.cell(row, 1).alignment = center
        ws.cell(row, 1).border    = border
        for m in range(1, 13):
            c = ws.cell(row, m + 1)
            c.fill = _day_fill(m, day); c.alignment = center; c.border = border

    # 月總計列
    ws.cell(34, 1).value     = "月總計/張"
    ws.cell(34, 1).font      = total_font
    ws.cell(34, 1).fill      = total_fill
    ws.cell(34, 1).alignment = center
    ws.cell(34, 1).border    = border
    for m in range(1, 13):
        c = ws.cell(34, m + 1)
        c.value = 0; c.font = total_font
        c.fill = total_fill; c.alignment = center; c.border = border

    # 年總計列（先設框線再合併）
    for c in range(1, 14):
        cell = ws.cell(35, c)
        cell.fill = grand_fill; cell.font = grand_font
        cell.alignment = center; cell.border = border
    ws.merge_cells("A35:M35")
    ws.cell(35, 1).value = "年總計/張：0"

    # 欄寬
    ws.column_dimensions["A"].width = 10
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 8

    return wb


def _sync_annual_to_monthly_wb(wb_m, year):
    """將折扣券統計.xlsx 的年度分頁複製到月份工作簿（同步最新資料）"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not os.path.exists(VOUCHER_STATS_FILE):
        return
    try:
        src_wb = openpyxl.load_workbook(VOUCHER_STATS_FILE, data_only=True)
    except Exception:
        return

    src_ws = src_wb.active
    annual_name = f"{year}年停車券統計"

    if annual_name in wb_m.sheetnames:
        del wb_m[annual_name]
    dst_ws = wb_m.create_sheet(annual_name)

    # 複製合併儲存格
    for rng in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(rng))

    # 複製欄寬
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width

    # 複製列高
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height

    # 複製儲存格內容與樣式（跳過合併次格）
    import copy as _copy
    from openpyxl.cell import MergedCell
    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            dst = dst_ws.cell(row=cell.row, column=cell.column)
            dst.value = cell.value
            if cell.font:
                dst.font = _copy.copy(cell.font)
            if cell.fill and cell.fill.patternType:
                dst.fill = _copy.copy(cell.fill)
            if cell.border:
                dst.border = _copy.copy(cell.border)
            if cell.alignment:
                dst.alignment = _copy.copy(cell.alignment)
            if cell.number_format:
                dst.number_format = cell.number_format

    # 合併列補框線：先解除合併 → 對整列補樣式 → 重新合併
    thin = Side(style="thin", color="000000")
    blk  = Border(left=thin, right=thin, top=thin, bottom=thin)
    merged_rows = {}
    for rng in list(dst_ws.merged_cells.ranges):
        if rng.min_row == rng.max_row:
            merged_rows[rng.min_row] = (rng.min_col, rng.max_col, str(rng))
    for row_num, (min_col, max_col, rng_str) in merged_rows.items():
        dst_ws.unmerge_cells(rng_str)
        ref = dst_ws.cell(row_num, min_col)
        for c in range(min_col, max_col + 1):
            cell = dst_ws.cell(row_num, c)
            cell.border = _copy.copy(blk)
            if c != min_col:
                cell.fill   = _copy.copy(ref.fill) if ref.fill else cell.fill
                cell.font   = _copy.copy(ref.font) if ref.font else cell.font
                cell.alignment = _copy.copy(ref.alignment) if ref.alignment else cell.alignment
        dst_ws.merge_cells(rng_str)


def rebuild_voucher_annual_in_monthly():
    """將年度折扣券統計分頁加入所有現有月份停車券 Excel"""
    import glob, time as _t, re
    files = [f for f in sorted(glob.glob(os.path.join(VOUCHER_STATS_DIR, "停車券????年??月.xlsx")))
             if not os.path.basename(f).startswith("~$")]
    if not files:
        print("無月份停車券檔案")
        return
    print(f"共 {len(files)} 個月份檔案，開始加入年度統計分頁...")
    tmp = r"C:\temp\parking\voucher_month_save.xlsx"
    for fpath in files:
        m = re.match(r".*停車券(\d{4})年\d{2}月\.xlsx", fpath)
        year = int(m.group(1)) if m else 2026
        try:
            wb = openpyxl.load_workbook(fpath)
        except Exception as e:
            print(f"  無法開啟 {os.path.basename(fpath)}: {e}")
            continue
        _sync_annual_to_monthly_wb(wb, year)
        # 年度統計置首
        annual_name = f"{year}年停車券統計"
        if annual_name in wb.sheetnames:
            wb.move_sheet(annual_name, offset=-wb.sheetnames.index(annual_name))
        wb.save(tmp)
        for _ in range(6):
            try:
                shutil.copy2(tmp, fpath)
                break
            except PermissionError:
                _t.sleep(10)
        print(f"  → 已更新：{os.path.basename(fpath)}")
    print("完成！")


def _update_voucher_stats(date_obj, count, rows=None):
    """更新折扣券年度統計 Excel：更新年度分頁 + 新增每日明細分頁"""
    import time as _time
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    year  = date_obj.year
    month = date_obj.month
    day   = date_obj.day
    date_str = date_obj.strftime("%Y-%m-%d")
    os.makedirs(VOUCHER_STATS_DIR, exist_ok=True)

    if os.path.exists(VOUCHER_STATS_FILE):
        wb = openpyxl.load_workbook(VOUCHER_STATS_FILE)
    else:
        wb = _create_voucher_stats_wb(year)

    # ── 年度統計分頁 ──────────────────────────────────────
    ws = wb.active
    data_row = day + 2
    col      = month + 1
    ws.cell(data_row, col).value = count if count > 0 else ""

    for m in range(1, 13):
        mc = m + 1
        monthly = sum(
            (ws.cell(r, mc).value or 0) for r in range(3, 34)
            if isinstance(ws.cell(r, mc).value, (int, float))
        )
        ws.cell(34, mc).value = monthly

    annual = sum(ws.cell(34, m + 1).value or 0 for m in range(1, 13))
    ws.cell(35, 1).value = f"年總計/張：{annual}"

    # ── 年度統計存檔 ──────────────────────────────────────
    tmp = r"C:\temp\parking\voucher_save.xlsx"
    wb.save(tmp)
    for _ in range(6):
        try:
            shutil.copy2(tmp, VOUCHER_STATS_FILE)
            break
        except PermissionError:
            _time.sleep(10)

    # ── 每日明細分頁（月份獨立檔案，與現金支付同樣結構）──────────
    if rows:
        from openpyxl.utils import get_column_letter as _gcl
        thin   = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        hdr_fill  = PatternFill("solid", fgColor="FFFFFF")
        hdr_font  = Font(bold=True, color="000000", size=11)
        total_font = Font(bold=True, color="FF0000")

        mpath = _voucher_monthly_path(date_obj)
        if os.path.exists(mpath):
            wb_m = openpyxl.load_workbook(mpath)
        else:
            wb_m = openpyxl.Workbook()
            if "Sheet" in wb_m.sheetnames:
                del wb_m["Sheet"]

        if date_str in wb_m.sheetnames:
            del wb_m[date_str]
        wd = wb_m.create_sheet(date_str)

        headers = ["車牌號碼", "入場時間", "出場時間", "支付金額", "電子優惠(分鐘)", "支付方式"]
        for ci2, h in enumerate(headers, start=1):
            c = wd.cell(1, ci2)
            c.value = h; c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = border

        total_min = 0
        for ri, (plate, entry, exit_, amt, disc, method) in enumerate(rows, start=2):
            wd.cell(ri, 1).value = plate
            wd.cell(ri, 2).value = entry
            wd.cell(ri, 3).value = exit_
            wd.cell(ri, 4).value = amt if amt else 0
            wd.cell(ri, 5).value = disc if disc else 0
            wd.cell(ri, 6).value = method
            try: total_min += float(disc) if disc else 0
            except: pass
            for ci2 in range(1, 7):
                wd.cell(ri, ci2).alignment = center
                wd.cell(ri, ci2).border    = border

        total_row = len(rows) + 2
        wd.cell(total_row, 3).value = f"{date_obj.strftime('%Y.%m.%d')}當日時數"
        wd.cell(total_row, 4).value = 0
        wd.cell(total_row, 5).value = int(total_min)
        for ci2 in range(1, 7):
            wd.cell(total_row, ci2).font      = total_font
            wd.cell(total_row, ci2).alignment = center
            wd.cell(total_row, ci2).border    = border

        def _cw(v):
            s = str(v) if v is not None else ""
            return sum(2 if '一' <= ch <= '鿿' or '＀' <= ch <= '￯' else 1 for ch in s)
        for ci2 in range(1, 7):
            max_w = max(_cw(wd.cell(r, ci2).value) for r in range(1, total_row + 1))
            wd.column_dimensions[_gcl(ci2)].width = max_w + 2

        # 加入年度統計分頁
        _sync_annual_to_monthly_wb(wb_m, year)

        # 依日期排序分頁（年度統計置首）
        date_sheets = sorted([s for s in wb_m.sheetnames
                               if len(s) == 10 and s[4] == '-' and s[7] == '-'])
        annual_name = f"{year}年停車券統計"
        if annual_name in wb_m.sheetnames:
            wb_m.move_sheet(annual_name, offset=-wb_m.sheetnames.index(annual_name))
        for idx, sn in enumerate(date_sheets):
            wb_m.move_sheet(sn, offset=idx + (1 if annual_name in wb_m.sheetnames else 0)
                            - wb_m.sheetnames.index(sn))

        tmp_m = r"C:\temp\parking\voucher_month_save.xlsx"
        wb_m.save(tmp_m)
        for _ in range(6):
            try:
                shutil.copy2(tmp_m, mpath)
                break
            except PermissionError:
                _time.sleep(10)

    print(f"  停車券統計：{date_str} {count}張")


def _sheet_to_png(excel_path, sheet_name, out_png):
    """將 Excel 分頁轉成 PNG（透過 Playwright 截圖）"""
    from playwright.sync_api import sync_playwright

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    # 取得合併儲存格範圍
    merged = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = (rng.min_row, rng.min_col,
                                  rng.max_row - rng.min_row + 1,
                                  rng.max_col - rng.min_col + 1)

    # 計算欄寬（像素）
    col_px = {}
    for c in range(1, ws.max_column + 1):
        from openpyxl.utils import get_column_letter
        ltr = get_column_letter(c)
        w = ws.column_dimensions[ltr].width or 8
        col_px[c] = max(int(w * 7), 40)

    html = ['<html><head><meta charset="utf-8"><style>',
            'body{font-family:Arial,sans-serif;font-size:11px;margin:8px;}',
            'table{border-collapse:collapse;}',
            'td{border:1px solid #bbb;padding:2px 5px;white-space:nowrap;text-align:center;}',
            '</style></head><body><table>']

    skip = set()
    for row in ws.iter_rows():
        html.append('<tr>')
        for cell in row:
            rc = (cell.row, cell.column)
            if rc in skip:
                continue
            val = cell.value if cell.value is not None else ''
            if isinstance(val, float) and val == int(val):
                val = int(val)
            # 貨幣格式
            if isinstance(val, (int, float)) and cell.number_format and '$' in str(cell.number_format):
                val = f'${val:,}'

            style = f'width:{col_px.get(cell.column, 60)}px;'
            if cell.font:
                if cell.font.bold:
                    style += 'font-weight:bold;'
                if cell.font.color and getattr(cell.font.color, 'type', '') == 'rgb':
                    rgb = cell.font.color.rgb
                    if rgb and rgb[-6:] not in ('000000',):
                        style += f'color:#{rgb[-6:]};'
            if cell.fill and cell.fill.patternType == 'solid':
                fg = getattr(cell.fill.fgColor, 'rgb', '')
                if fg and fg[-6:] not in ('000000', 'FFFFFF'):
                    style += f'background:#{fg[-6:]};'

            colspan = rowspan = 1
            if rc in merged:
                mr, mc, rs, cs = merged[rc]
                if (mr, mc) == rc:
                    rowspan, colspan = rs, cs
                    for dr in range(rs):
                        for dc in range(cs):
                            if (dr, dc) != (0, 0):
                                skip.add((mr + dr, mc + dc))
                else:
                    continue

            attrs = f'style="{style}"'
            if colspan > 1: attrs += f' colspan="{colspan}"'
            if rowspan > 1: attrs += f' rowspan="{rowspan}"'
            html.append(f'<td {attrs}>{val}</td>')
        html.append('</tr>')
    html.append('</table></body></html>')

    tmp_html = r'C:\temp\parking\tmp_sheet.html'
    with open(tmp_html, 'w', encoding='utf-8') as f:
        f.write('\n'.join(html))

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(viewport={'width': 1400, 'height': 900})
        page.goto(f'file:///{tmp_html}')
        page.locator('table').screenshot(path=out_png)
        browser.close()

    return out_png


def _send_report_email(date_obj, png_files):
    """寄送每日報表圖檔至 mikulai1114@gmail.com"""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage

    if not os.path.exists(GMAIL_CONFIG):
        print(f'  [郵件] 找不到 {GMAIL_CONFIG}，略過寄信')
        print(f'  [郵件] 請建立該檔，第1行：Gmail帳號，第2行：應用程式密碼')
        return

    with open(GMAIL_CONFIG, encoding='utf-8') as f:
        lines = [l.strip() for l in f.readlines()]
    if len(lines) < 2:
        print('  [郵件] gmail_config.txt 格式錯誤')
        return

    sender = lines[0]
    app_pw = lines[1]
    to_addr = 'mikulai1114@gmail.com'
    date_str = date_obj.strftime('%Y-%m-%d')

    msg = MIMEMultipart()
    msg['From']    = sender
    msg['To']      = to_addr
    msg['Subject'] = f'磐鈺雲華停車場 {date_str} 每日報表'
    msg.attach(MIMEText(
        f'附件為 {date_str} 每日停車分頁及年度統計分頁截圖。\n\n磐鈺雲華商業停車場 自動報表系統',
        'plain', 'utf-8'))

    for png in png_files:
        if os.path.exists(png):
            with open(png, 'rb') as f:
                img = MIMEImage(f.read())
                img.add_header('Content-Disposition', 'attachment',
                               filename=os.path.basename(png))
                msg.attach(img)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(sender, app_pw)
            smtp.sendmail(sender, to_addr, msg.as_string())
        print(f'  [郵件] 已寄出至 {to_addr}')
    except Exception as e:
        print(f'  [郵件] 寄送失敗：{e}')


def clean_invoice_sections():
    """移除所有每日分頁中的電子發票明細區塊"""
    import glob, time
    xlsx_files = [f for f in sorted(glob.glob(os.path.join(MONTHLY_DIR, "*.xlsx")))
                  if not os.path.basename(f).startswith("~$")]
    tmp = r"C:\temp\parking\month_save.xlsx"

    for mpath in xlsx_files:
        wb = openpyxl.load_workbook(mpath)
        changed = False
        for sn in wb.sheetnames:
            try:
                datetime.strptime(sn, "%Y-%m-%d")
            except ValueError:
                continue
            ws = wb[sn]
            inv_start = None
            for r in range(1, ws.max_row + 1):
                v1 = ws.cell(r, 1).value
                v2 = ws.cell(r, 2).value
                if v1 in ("電子發票明細", "發票日期") or v2 in ("電子發票明細", "發票日期", "品名"):
                    inv_start = r
                    break
            if inv_start is None:
                continue
            # 若 inv_start 上一列是空白列則一併刪除，否則從 inv_start 開始刪
            prev_row = inv_start - 1
            if prev_row >= 1 and all(ws.cell(prev_row, c).value is None for c in range(1, 10)):
                del_from = prev_row
            else:
                del_from = inv_start
            ws.delete_rows(del_from, ws.max_row - del_from + 1)
            changed = True
            print(f"  {sn}: 已移除")

        if changed:
            wb.save(tmp)
            for _ in range(6):
                try:
                    shutil.copy2(tmp, mpath)
                    break
                except PermissionError:
                    time.sleep(10)
            print(f"  → 已儲存 {os.path.basename(mpath)}")

    print("清除完成！")


def rebuild_invoices():
    """從電子發票平台抓取所有有資料的日期，寫入月份 Excel 每日分頁"""
    import glob, time as _time
    from playwright.sync_api import sync_playwright as _sp

    # 找出所有已有每日分頁的日期
    xlsx_files = [f for f in sorted(glob.glob(os.path.join(MONTHLY_DIR, "*.xlsx")))
                  if not os.path.basename(f).startswith("~$")]
    date_map = {}  # date_str → (monthly_path, sheet_name)
    for path in xlsx_files:
        wb = openpyxl.load_workbook(path, read_only=True)
        for sn in wb.sheetnames:
            try:
                datetime.strptime(sn, "%Y-%m-%d")
                date_map[sn] = path
            except ValueError:
                pass
        wb.close()

    print(f"找到 {len(date_map)} 個每日分頁，開始查詢電子發票...\n")

    with _sp() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        print("登入電子發票平台...")
        if not _inv_login(page):
            print("登入失敗！")
            browser.close()
            return
        print("登入成功\n")
        page.on("dialog", lambda d: d.accept())

        results = {}  # date_str → invoice_rows
        for date_str in sorted(date_map.keys()):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            inv_path = _inv_fetch_excel(page, date_obj)
            if inv_path:
                rows = _inv_parse_excel(inv_path)
                if rows:
                    results[date_str] = rows
                    print(f"  {date_str}: {len(rows)} 張發票")
                else:
                    print(f"  {date_str}: 無有效發票資料")
            else:
                print(f"  {date_str}: 查無資料")

        browser.close()

    if not results:
        print("\n無任何發票資料")
        return

    print(f"\n共 {len(results)} 天有發票資料，寫入月份 Excel...")
    tmp = r"C:\temp\parking\month_save.xlsx"

    # 依月份分組處理
    monthly_paths = set(date_map[d] for d in results)
    for mpath in sorted(monthly_paths):
        wb = openpyxl.load_workbook(mpath)
        changed = False
        for date_str, inv_rows in results.items():
            if date_map[date_str] != mpath:
                continue
            if date_str not in wb.sheetnames:
                continue
            ws = wb[date_str]
            # 確認尚未寫過（避免重複，標題在 col2）
            body_vals = []
            for r in range(max(1, ws.max_row - 5), ws.max_row + 1):
                for c in (1, 2):
                    v = ws.cell(r, c).value
                    if v:
                        body_vals.append(v)
            if "電子發票明細" in body_vals:
                print(f"  {date_str}: 已有發票區塊，略過")
                continue
            _append_invoice_section(ws, inv_rows, datetime.strptime(date_str, "%Y-%m-%d"))
            _auto_col_width(ws)
            changed = True
            print(f"  {date_str}: 已寫入 {len(inv_rows)} 張")

        if changed:
            import time
            wb.save(tmp)
            for _ in range(6):
                try:
                    shutil.copy2(tmp, mpath)
                    break
                except PermissionError:
                    time.sleep(10)
            print(f"  → 已儲存 {os.path.basename(mpath)}")

    print("\n完成！")

    # 發票寫入後，重建所有月份的年度收入統計分頁
    rebuild_stats()


def rebuild_stats():
    """重建月份資料夾內所有 Excel 的年度收入統計分頁"""
    import glob, time, re
    all_monthly = [f for f in sorted(glob.glob(os.path.join(MONTHLY_DIR, "*.xlsx")))
                   if not os.path.basename(f).startswith("~$")]
    if not all_monthly:
        print("月份資料夾內無 Excel 檔案")
        return
    tmp = r"C:\temp\parking\month_save.xlsx"
    print(f"重建年度收入統計分頁（共 {len(all_monthly)} 個月份檔案）...")
    for mpath in all_monthly:
        try:
            wb = openpyxl.load_workbook(mpath)
        except Exception as e:
            print(f"  ⚠ 無法開啟 {os.path.basename(mpath)}: {e}")
            continue
        fname = os.path.basename(mpath)
        m = re.match(r"(\d{4})年(\d{2})月\.xlsx", fname)
        if not m:
            continue
        year  = int(m.group(1))
        month = int(m.group(2))
        _build_stats_sheet(wb, year, month)
        wb.save(tmp)
        for _ in range(6):
            try:
                shutil.copy2(tmp, mpath)
                break
            except PermissionError:
                time.sleep(10)
        print(f"  → 已更新：{fname}")
    _rebuild_invoice_record()
    print("年度統計重建完成！")


def run_batch_voucher(start_date, end_date):
    """批次補跑眾陽時數券資料（登入一次，逐日下載）"""
    d   = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date,   "%Y-%m-%d")
    dates = []
    while d <= end:
        dates.append(d)
        d += timedelta(days=1)

    total_days = len(dates)
    print(f"批次補跑停車券：{start_date} ~ {end_date}，共 {total_days} 天\n")
    os.makedirs(DL_DIR, exist_ok=True)
    os.makedirs(VOUCHER_STATS_DIR, exist_ok=True)

    errors = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        print("登入中...")
        frame = _login_and_open(page)
        print("登入成功，開始逐日補跑\n")

        for i, date_obj in enumerate(dates):
            target = date_obj.strftime("%Y-%m-%d")
            print(f"[{i+1:3}/{total_days}] {target}", end="  ", flush=True)

            voucher_path = os.path.join(DL_DIR, f"voucher_{target}.xlsx")
            try:
                _set_filter_voucher(frame, page,
                                    f"{target} 00:00:00",
                                    f"{target} 23:59:59")
                _search(frame, page)

                row_count = frame.evaluate("""
                    (function(){
                        var rows = document.querySelectorAll('table tbody tr');
                        if(rows.length === 0) return 0;
                        var txt = (rows[0].innerText || '').trim();
                        if(txt.indexOf('無數據') >= 0 ||
                           txt.indexOf('暫無數據') >= 0 ||
                           txt.indexOf('No data') >= 0) return 0;
                        return rows.length;
                    })();
                """)

                if row_count == 0:
                    print("→ 無資料")
                    _update_voucher_stats(date_obj, 0, [])
                else:
                    with page.expect_download(timeout=20000) as dl_info:
                        frame.evaluate("""
                            (function(){
                                var btns=document.querySelectorAll('button,.layui-btn,a');
                                for(var btn of btns){
                                    if((btn.innerText||'').trim().includes('Excel')){btn.click();break;}
                                }
                            })();
                        """)
                    dl_info.value.save_as(voucher_path)
                    v_count, v_rows = _parse_voucher_excel(voucher_path)
                    _update_voucher_stats(date_obj, v_count, v_rows)
                    print(f"→ {v_count} 張  {sum(float(r[4]) if r[4] else 0 for r in v_rows):.0f} 分鐘")

            except Exception as e:
                msg = str(e)[:60]
                print(f"→ 略過（{msg}）")
                errors.append((target, msg))

        browser.close()

    print(f"\n批次完成！共 {total_days} 天，失敗 {len(errors)} 筆")
    if errors:
        print("失敗清單：")
        for d2, e in errors:
            print(f"  {d2}: {e}")


if __name__ == "__main__":
    if len(sys.argv) >= 2 and sys.argv[1] == "rebuild":
        rebuild_monthly()
    elif len(sys.argv) >= 2 and sys.argv[1] == "invoices":
        rebuild_invoices()
    elif len(sys.argv) >= 2 and sys.argv[1] == "clean_inv":
        clean_invoice_sections()
    elif len(sys.argv) >= 2 and sys.argv[1] == "rebuild_stats":
        rebuild_stats()
    elif len(sys.argv) >= 2 and sys.argv[1] == "voucher_annual":
        rebuild_voucher_annual_in_monthly()
    elif len(sys.argv) >= 2 and sys.argv[1] == "voucher":
        if len(sys.argv) >= 4:
            run_batch_voucher(sys.argv[2], sys.argv[3])
        else:
            yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
            run_batch_voucher("2026-01-01", yesterday)
    elif len(sys.argv) >= 3:
        run_batch(sys.argv[1], sys.argv[2])
    else:
        run(sys.argv[1] if len(sys.argv) > 1 else None)
