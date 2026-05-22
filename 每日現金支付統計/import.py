import sys, os, shutil
from datetime import datetime, timedelta
from playwright.sync_api import sync_playwright
import openpyxl

LOGIN_URL = "http://125.227.64.94:48140/Account/Login"
USERNAME  = "23208598"
PASSWORD  = "23208598"
DL_DIR    = r"C:\temp\parking"
CONFIG    = r"C:\temp\parking\config.txt"
TMP_PATH  = r"C:\temp\parking\wb_save.xlsx"

MONTHS = ["一月","二月","三月","四月","五月",
          "六月","七月","八月","九月","十月",
          "十一月","十二月"]
MONTHS_NUM = ["01","02","03","04","05","06",
              "07","08","09","10","11","12"]
WEEKDAY_ZH = ["一","二","三","四","五","六","日"]

with open(CONFIG, encoding="utf-8") as _f:
    WORKBOOK = _f.read().strip()

# 月份 Excel 存放資料夾（與主工作簿同層）
MONTHLY_DIR = os.path.join(os.path.dirname(WORKBOOK), "月份資料")


def _parse_source(excel_path):
    src = openpyxl.load_workbook(excel_path)
    ws  = src.active
    headers = [str(c.value) if c.value else "" for c in ws[1]]

    def ci(names):
        for n in names:
            if n in headers: return headers.index(n)
        return None

    i_plate   = ci(["車牌號碼","車牌"])
    i_entry   = ci(["入場時間","入場"])
    i_exit    = ci(["出場時間","出場"])
    i_amount  = ci(["支付金額","金額"])
    i_method  = ci(["支付方式"])
    i_invoice = ci(["發票號碼"])
    i_carrier = ci(["統編/載具","載具"])

    rows_data, total = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row): continue
        r = list(row)
        plate_val = str(r[i_plate]).strip() if i_plate is not None and r[i_plate] else ""
        if plate_val in ("總計", "合計", "小計", ""):
            continue
        try:
            amt = float(r[i_amount]) if i_amount is not None and r[i_amount] else 0
        except:
            amt = 0
        total += amt
        rows_data.append(r)

    cols = (i_plate, i_entry, i_exit, i_amount, i_method, i_invoice, i_carrier)
    return rows_data, total, cols


def _monthly_excel_path(date_obj):
    """月份 Excel 檔案路徑"""
    fname = f"{date_obj.year}年{MONTHS_NUM[date_obj.month-1]}月.xlsx"
    return os.path.join(MONTHLY_DIR, fname)


def _get_or_create_monthly_wb(monthly_wbs, date_obj):
    """取得或建立月份工作簿（in-memory）"""
    key = (date_obj.year, date_obj.month)
    if key not in monthly_wbs:
        path = _monthly_excel_path(date_obj)
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            # 移除舊版的單一合併工作表（如果存在）
            month_title = MONTHS[date_obj.month - 1]
            if month_title in wb.sheetnames and len(wb.sheetnames) == 1:
                del wb[month_title]
        else:
            wb = openpyxl.Workbook()
            # 移除預設空白工作表
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        monthly_wbs[key] = wb
    return monthly_wbs[key]


def _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols):
    """每天建立獨立分頁，分頁名稱為日期"""
    wb       = _get_or_create_monthly_wb(monthly_wbs, date_obj)
    sheet_name = date_obj.strftime("%Y-%m-%d")

    # 若已有同名分頁則先刪除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # 表頭（7欄，不含日期欄）
    for ci2, h in enumerate(
        ["車牌號碼","進場時間","出場時間","支付金額","支付方式","發票號碼","載具"],
        start=1
    ):
        ws.cell(1, ci2).value = h

    # 寫入資料
    (i_plate, i_entry, i_exit, i_amount, i_method, i_invoice, i_carrier) = cols
    src_cols = [i_plate, i_entry, i_exit, i_amount, i_method, i_invoice, i_carrier]
    for ri, rd in enumerate(rows_data, start=2):
        for ci2, sc in enumerate(src_cols, start=1):
            if sc is not None:
                ws.cell(ri, ci2).value = rd[sc]

    # 合計列在最下方
    total_row = len(rows_data) + 2
    ws.cell(total_row, 1).value = "合計"
    ws.cell(total_row, 3).value = date_obj.strftime("%Y/%m/%d")
    ws.cell(total_row, 4).value = int(total)


def _display_width(text):
    """計算顯示寬度：中文字符算 2，英數算 1"""
    w = 0
    for ch in str(text):
        w += 2 if ord(ch) > 127 else 1
    return w


def _auto_col_width(ws):
    """依內容自動調整欄寬（含中文寬度）並加框線"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                w = _display_width(cell.value)
                if w > col_widths.get(cell.column, 0):
                    col_widths[cell.column] = w
            cell.border = border

    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 45)


def _build_stats_sheet(wb, year, month):
    """建立年度收入統計分頁（複製自主工作簿，值模式，放置於最前）"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sheet_name = "年度收入統計"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)   # 最前

    # 讀取主工作簿年度統計（data_only=True 取計算值）
    main_wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    src_name = None
    for name in main_wb.sheetnames:
        if "年度" in name and "統計" in name:
            src_name = name
            break
    if not src_name:
        return
    src_ws = main_wb[src_name]

    # 樣式
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill  = PatternFill("solid", fgColor="1F4E79")
    title_font  = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill  = PatternFill("solid", fgColor="BDD7EE")
    total_font  = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            src_cell  = src_ws.cell(r, c)
            dest_cell = ws.cell(r, c)
            dest_cell.value = src_cell.value
            dest_cell.border = border

            # 標題列
            if r == 1:
                dest_cell.font      = title_font
                dest_cell.fill      = title_fill
                dest_cell.alignment = Alignment(horizontal="center", vertical="center")
            # 表頭列
            elif r == 2:
                dest_cell.font      = header_font
                dest_cell.fill      = header_fill
                dest_cell.alignment = center
            # 小計／總計列
            elif src_cell.value in ("小計", "總計"):
                dest_cell.font      = total_font
                dest_cell.fill      = total_fill
                dest_cell.alignment = center if c == 1 else right
            # 數值欄（第2欄以後）
            elif c > 1 and isinstance(src_cell.value, (int, float)):
                dest_cell.alignment = right
            # 第1欄（日期序號）
            elif c == 1:
                dest_cell.alignment = center

    # 合併標題列（同原始）
    try:
        ws.merge_cells(f"A1:{get_column_letter(src_ws.max_column)}1")
    except Exception:
        pass

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # 欄寬
    ws.column_dimensions["A"].width = 8
    for c in range(2, src_ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


def _build_nav_sheet(wb):
    """建立或更新「選擇日期」導覽分頁（下拉選單 + 超連結）"""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    nav_name = "選擇日期"
    # 取得所有日期分頁（排除導覽分頁本身）
    date_sheets = [s for s in wb.sheetnames if s != nav_name]

    # 刪除舊導覽分頁，重建置於最前
    if nav_name in wb.sheetnames:
        del wb[nav_name]
    ws = wb.create_sheet(nav_name, 0)

    # 樣式
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    label_font  = Font(bold=True, size=11)
    link_font   = Font(color="0000FF", underline="single", size=11)

    # 標題列
    ws.merge_cells("A1:C1")
    ws["A1"].value     = "請選擇日期後點擊「前往」"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 標籤
    ws["A2"].value = "選擇日期："
    ws["A2"].font  = label_font
    ws["A2"].border = border

    # 下拉選單儲存格
    ws["B2"].value     = date_sheets[0] if date_sheets else ""
    ws["B2"].font      = Font(size=11)
    ws["B2"].border    = border
    ws["B2"].alignment = Alignment(horizontal="center")

    # 資料驗證：下拉清單
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(date_sheets) + '"',
        allow_blank=False,
        showDropDown=False
    )
    ws.add_data_validation(dv)
    dv.add(ws["B2"])

    # 超連結「前往」按鈕（HYPERLINK 公式）
    ws["C2"].value     = '=HYPERLINK("#\'"&B2&"\'!A1","▶ 前往")'
    ws["C2"].font      = link_font
    ws["C2"].border    = border
    ws["C2"].alignment = Alignment(horizontal="center")

    # 下方列出所有日期的快速連結
    ws["A4"].value = "所有日期快速連結："
    ws["A4"].font  = Font(bold=True, size=10)

    for i, sheet in enumerate(date_sheets):
        row = i + 5
        ws.cell(row, 1).value     = f'=HYPERLINK("#\'{sheet}\'!A1","{sheet}")'
        ws.cell(row, 1).font      = link_font
        ws.cell(row, 1).border    = border
        ws.cell(row, 1).alignment = Alignment(horizontal="center")

    # 欄寬
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12


def _save_monthly_wbs(monthly_wbs):
    os.makedirs(MONTHLY_DIR, exist_ok=True)
    tmp = r"C:\temp\parking\month_save.xlsx"
    for (year, month), wb in monthly_wbs.items():
        for ws in wb.worksheets:
            _auto_col_width(ws)
        _build_nav_sheet(wb)          # 選擇日期（第2分頁）
        _build_stats_sheet(wb, year, month)  # 月份收入統計（第1分頁）
        key_obj = datetime(year, month, 1)
        path = _monthly_excel_path(key_obj)
        wb.save(tmp)
        shutil.copy2(tmp, path)


def _write_monthly(wb, month_name, day, total):
    wb[month_name].cell(day + 2, 4).value = int(total)


def _write_daily_summary(wb, date_obj, rows_data, total, i_amount):
    date_str    = date_obj.strftime("%Y/%m/%d")
    weekday_zh  = WEEKDAY_ZH[date_obj.weekday()]
    type_str    = "假日" if date_obj.weekday() >= 5 else "平日"
    total_trips = len(rows_data)
    paid_trips  = sum(1 for rd in rows_data
                      if (float(rd[i_amount]) if i_amount is not None and rd[i_amount] else 0) > 0)
    free_trips  = total_trips - paid_trips
    cash_income = int(total)

    ws_d = wb["每日彙整"]
    target_row = None
    for row_idx in range(2, ws_d.max_row + 1):
        val = ws_d.cell(row_idx, 1).value
        if val is not None and str(val).strip() == date_str:
            target_row = row_idx
            break
    if target_row is None:
        ws_d.insert_rows(2)
        target_row = 2

    ws_d.cell(target_row, 1).value = date_str
    ws_d.cell(target_row, 2).value = weekday_zh
    ws_d.cell(target_row, 3).value = type_str
    ws_d.cell(target_row, 4).value = total_trips
    ws_d.cell(target_row, 5).value = paid_trips
    ws_d.cell(target_row, 6).value = free_trips
    ws_d.cell(target_row, 7).value = cash_income
    ws_d.cell(target_row, 8).value = 0
    ws_d.cell(target_row, 9).value = cash_income
    return total_trips, cash_income


def _save_wb(wb):
    import time
    wb.save(TMP_PATH)
    for attempt in range(6):
        try:
            shutil.copy2(TMP_PATH, WORKBOOK)
            return
        except PermissionError:
            if attempt == 0:
                print("\n  ⚠ 主檔被 Excel 開啟，等待 10 秒後重試...")
            time.sleep(10)
    raise PermissionError("無法寫入主檔，請關閉 Excel 後重試")


def _set_filter(frame, page, start_str, end_str):
    frame.evaluate(f"""
        (function(){{
            var inputs = document.querySelectorAll('input.layui-input, input[type=text]');
            if(inputs[0]){{ inputs[0].value='{start_str}'; inputs[0].dispatchEvent(new Event('change',{{bubbles:true}})); }}
            if(inputs[1]){{ inputs[1].value='{end_str}';   inputs[1].dispatchEvent(new Event('change',{{bubbles:true}})); }}
            var selects = document.querySelectorAll('select');
            for(var sel of selects){{
                for(var opt of sel.options){{
                    if(opt.text && opt.text.includes('現金支付')){{
                        sel.value=opt.value; sel.dispatchEvent(new Event('change',{{bubbles:true}})); break;
                    }}
                }}
            }}
        }})();
    """)
    page.wait_for_timeout(800)


def _search(frame, page):
    frame.evaluate("""
        (function(){
            var btns=document.querySelectorAll('button,.layui-btn');
            for(var btn of btns){
                if((btn.innerText||'').replace(/\\s+/g,'').includes('搜索')){btn.click();break;}
            }
        })();
    """)
    page.wait_for_load_state("networkidle", timeout=20000)
    page.wait_for_timeout(1500)


def _export(frame, page, target):
    save_path = os.path.join(DL_DIR, f"parking_{target}.xlsx")
    with page.expect_download(timeout=20000) as dl_info:
        frame.evaluate("""
            (function(){
                var btns=document.querySelectorAll('button,.layui-btn,a');
                for(var btn of btns){
                    if((btn.innerText||'').trim().includes('Excel')){btn.click();break;}
                }
            })();
        """)
    dl_info.value.save_as(save_path)
    return save_path


def _login_and_open(page):
    page.goto(LOGIN_URL, timeout=30000)
    page.fill("input[name=username]", USERNAME)
    page.fill("input[name=password]", PASSWORD)
    page.keyboard.press("Enter")
    page.wait_for_load_state("networkidle", timeout=20000)

    nav = page.locator(".layui-nav-item")
    nav.nth(2).click()
    page.wait_for_timeout(1000)
    nav.nth(2).locator("dd a, .layui-nav-child a").first.click()

    frame = None
    for _ in range(10):
        page.wait_for_timeout(1000)
        for f in page.frames:
            if "ChargeRecord" in f.url:
                frame = f
                break
        if frame:
            break
    if not frame:
        raise Exception("找不到支付流水頁面")
    return frame


# ── 單日匯入 ──────────────────────────────────────────────
def run(target=None):
    if target is None:
        target = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    date_obj   = datetime.strptime(target, "%Y-%m-%d")
    month_name = MONTHS[date_obj.month - 1]
    day        = date_obj.day
    os.makedirs(DL_DIR, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        print(f"[1] 登入 ({target})...")
        frame = _login_and_open(page)
        print(f"    OK")

        print(f"[2] 設定日期 + 現金支付...")
        _set_filter(frame, page, f"{target} 00:00:00", f"{target} 23:59:59")

        print("[3] 搜索...")
        _search(frame, page)

        print("[4] 匯出 Excel...")
        save_path = _export(frame, page, target)
        browser.close()

    rows_data, total, cols = _parse_source(save_path)
    print(f"  筆數:{len(rows_data)}  合計:${int(total):,}")

    # 主工作簿
    wb = openpyxl.load_workbook(WORKBOOK)
    _write_monthly(wb, month_name, day, total)
    _write_daily_summary(wb, date_obj, rows_data, total, cols[3])
    _save_wb(wb)

    # 月份 Excel
    monthly_wbs = {}
    _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols)
    _save_monthly_wbs(monthly_wbs)

    print(f"  已寫入：{month_name} 第{day}天 = ${int(total):,}")
    print("完成！")


# ── 批次匯入 ──────────────────────────────────────────────
def run_batch(start_date, end_date):
    d   = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date,   "%Y-%m-%d")
    dates = []
    while d <= end:
        dates.append(d)
        d += timedelta(days=1)

    total_days = len(dates)
    print(f"批次匯入：{start_date} ~ {end_date}，共 {total_days} 天\n")
    os.makedirs(DL_DIR, exist_ok=True)
    os.makedirs(MONTHLY_DIR, exist_ok=True)

    wb          = openpyxl.load_workbook(WORKBOOK)
    monthly_wbs = {}   # (year, month) → workbook
    errors      = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        print("登入中...")
        frame = _login_and_open(page)
        print("登入成功，開始逐日匯入\n")

        for i, date_obj in enumerate(dates):
            target     = date_obj.strftime("%Y-%m-%d")
            month_name = MONTHS[date_obj.month - 1]
            day        = date_obj.day

            print(f"[{i+1:3}/{total_days}] {target}", end="  ", flush=True)

            try:
                _set_filter(frame, page,
                            f"{target} 00:00:00",
                            f"{target} 23:59:59")
                _search(frame, page)
                save_path = _export(frame, page, target)
                rows_data, total, cols = _parse_source(save_path)

                _write_monthly(wb, month_name, day, total)
                _write_daily_summary(wb, date_obj, rows_data, total, cols[3])
                _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols)

                print(f"→ {len(rows_data):3}筆  ${int(total):>7,}")
            except Exception as e:
                msg = str(e)[:60]
                print(f"→ 略過（{msg}）")
                errors.append((target, msg))

            # 每 10 天存一次
            if (i + 1) % 10 == 0:
                _save_wb(wb)
                _save_monthly_wbs(monthly_wbs)
                print(f"         ↳ 已存檔 ({i+1}/{total_days})")

        browser.close()

    _save_wb(wb)
    _save_monthly_wbs(monthly_wbs)
    print(f"\n批次完成！共 {total_days} 天，失敗 {len(errors)} 筆")
    if errors:
        print("失敗清單：")
        for d, e in errors:
            print(f"  {d}: {e}")


def rebuild_monthly():
    """從本機已下載的 Excel 重建月份工作簿（每天一個分頁）"""
    import glob
    files = sorted(glob.glob(os.path.join(DL_DIR, "parking_*.xlsx")))
    print(f"找到 {len(files)} 個每日 Excel，開始重建月份工作簿...\n")

    monthly_wbs = {}
    for path in files:
        fname = os.path.basename(path)          # parking_2026-01-01.xlsx
        date_str = fname[8:18]                  # 2026-01-01
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            continue

        rows_data, total, cols = _parse_source(path)
        _append_to_monthly(monthly_wbs, date_obj, rows_data, total, cols)
        print(f"  {date_str}  {len(rows_data):3}筆  ${int(total):>7,}")

    # 存檔
    import time
    os.makedirs(MONTHLY_DIR, exist_ok=True)
    tmp = r"C:\temp\parking\month_save.xlsx"
    for (year, month), wb in monthly_wbs.items():
        for ws in wb.worksheets:
            _auto_col_width(ws)
        _build_nav_sheet(wb)
        _build_stats_sheet(wb, year, month)
        key_obj = datetime(year, month, 1)
        dest = _monthly_excel_path(key_obj)
        wb.save(tmp)
        for attempt in range(6):
            try:
                shutil.copy2(tmp, dest)
                break
            except PermissionError:
                print(f"  ⚠ {os.path.basename(dest)} 被開啟中，等待 10 秒...")
                time.sleep(10)
        print(f"\n已儲存：{os.path.basename(dest)}（{len(wb.sheetnames)} 個分頁）")

    print("\n重建完成！")


if __name__ == "__main__":
    if len(sys.argv) >= 2 and sys.argv[1] == "rebuild":
        rebuild_monthly()
    elif len(sys.argv) >= 3:
        run_batch(sys.argv[1], sys.argv[2])
    else:
        run(sys.argv[1] if len(sys.argv) > 1 else None)
