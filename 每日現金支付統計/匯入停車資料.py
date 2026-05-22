import sys, os
from datetime import datetime, timedelta
from playwright.sync_api import sync_playwright
import openpyxl

LOGIN_URL = "http://125.227.64.94:48140/Account/Login"
USERNAME  = "23208598"
PASSWORD  = "23208598"
DL_DIR    = r"C:\temp\parking"
CONFIG    = r"C:\temp\parking\config.txt"

# 一月=一月 二月=二月 ... 十一月=十一月 十二月=十二月
MONTHS = ["一月","二月","三月","四月","五月",
          "六月","七月","八月","九月","十月",
          "十一月","十二月"]

with open(CONFIG, encoding="utf-8") as _f:
    WORKBOOK = _f.read().strip()

def run(target=None):
    if target is None:
        target = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    date_obj   = datetime.strptime(target, "%Y-%m-%d")
    start_str  = f"{target} 00:00:00"
    end_str    = f"{target} 23:59:59"
    month_name = MONTHS[date_obj.month - 1]
    day        = date_obj.day
    os.makedirs(DL_DIR, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        print(f"[1/5] Login ({target})...")
        page.goto(LOGIN_URL, timeout=30000)
        page.fill("input[name=username]", USERNAME)
        page.fill("input[name=password]", PASSWORD)
        page.keyboard.press("Enter")
        page.wait_for_load_state("networkidle", timeout=20000)
        print("    OK")

        print("[2/5] Open payment page...")
        nav = page.locator(".layui-nav-item")
        nav.nth(2).click()
        page.wait_for_timeout(1000)
        nav.nth(2).locator("dd a, .layui-nav-child a").first.click()

        frame = None
        for _ in range(10):
            page.wait_for_timeout(1000)
            for f in page.frames:
                if "ChargeRecord" in f.url:
                    frame = f
                    break
            if frame:
                break
        if not frame:
            raise Exception("Cannot find payment page frame")
        print(f"    frame: {frame.url}")

        print(f"[3/5] Set date {target} + cash...")
        frame.wait_for_load_state("domcontentloaded", timeout=10000)
        frame.evaluate(f"""
            (function(){{
                var inputs = document.querySelectorAll('input.layui-input, input[type=text]');
                if(inputs[0]){{ inputs[0].value='{start_str}'; inputs[0].dispatchEvent(new Event('change',{{bubbles:true}})); }}
                if(inputs[1]){{ inputs[1].value='{end_str}';   inputs[1].dispatchEvent(new Event('change',{{bubbles:true}})); }}
                var selects = document.querySelectorAll('select');
                for(var sel of selects){{
                    for(var opt of sel.options){{
                        if(opt.text && opt.text.includes('現金支付')){{
                            sel.value=opt.value; sel.dispatchEvent(new Event('change',{{bubbles:true}})); break;
                        }}
                    }}
                }}
            }})();
        """)
        page.wait_for_timeout(800)

        print("[4/5] Search...")
        frame.evaluate("""
            (function(){
                var btns=document.querySelectorAll('button,.layui-btn');
                for(var btn of btns){
                    if((btn.innerText||'').replace(/\\s+/g,'').includes('搜索')){btn.click();break;}
                }
            })();
        """)
        page.wait_for_load_state("networkidle", timeout=20000)
        page.wait_for_timeout(2000)
        print("    Done")

        print("[5/5] Export Excel...")
        with page.expect_download(timeout=30000) as dl_info:
            frame.evaluate("""
                (function(){
                    var btns=document.querySelectorAll('button,.layui-btn,a');
                    for(var btn of btns){
                        if((btn.innerText||'').trim().includes('Excel')){btn.click();break;}
                    }
                })();
            """)
        dl = dl_info.value
        save_path = os.path.join(DL_DIR, f"parking_{target}.xlsx")
        dl.save_as(save_path)
        print(f"    Saved: {save_path}")
        browser.close()

    process(save_path, month_name, day, target)

def process(excel_path, month_name, day, target):
    print("\nProcessing...")
    src     = openpyxl.load_workbook(excel_path)
    ws_src  = src.active
    headers = [str(c.value) if c.value else "" for c in ws_src[1]]

    def ci(names):
        for n in names:
            if n in headers: return headers.index(n)
        return None

    # 車牌號碼=車牌號碼 入場時間=入場時間 etc.
    i_plate   = ci(["車牌號碼","車牌"])
    i_entry   = ci(["入場時間","入場"])
    i_exit    = ci(["出場時間","出場"])
    i_amount  = ci(["支付金額","金額"])
    i_method  = ci(["支付方式"])
    i_invoice = ci(["發票號碼"])
    i_carrier = ci(["統編/載具","載具"])

    rows_data, total = [], 0
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row): continue
        r = list(row)
        try:
            amt = float(r[i_amount]) if i_amount is not None and r[i_amount] else 0
        except:
            amt = 0
        total += amt
        rows_data.append(r)
    print(f"  Records:{len(rows_data)}  Total:${int(total):,}")

    wb = openpyxl.load_workbook(WORKBOOK)
    if target in wb.sheetnames:
        del wb[target]
    ws_raw = wb.create_sheet(target)
    headers = ["車牌號碼","進場時間","出場時間","支付金額","支付方式","發票號碼","載具"]
    for ci2, h in enumerate(headers, start=1):
        ws_raw.cell(1, ci2).value = h
    cols = [i_plate, i_entry, i_exit, i_amount, i_method, i_invoice, i_carrier]
    for ri, rd in enumerate(rows_data, start=2):
        for ci2, sc in enumerate(cols, start=1):
            if sc is not None: ws_raw.cell(ri, ci2).value = rd[sc]

    ws_m = wb[month_name]
    ws_m.cell(day + 2, 4).value = int(total)
    wb.save(WORKBOOK)
    print(f"  Saved: {month_name} day {day} = ${int(total):,}")
    print("Complete!")

if __name__ == "__main__":
    t = sys.argv[1] if len(sys.argv) > 1 else None
    run(t)
